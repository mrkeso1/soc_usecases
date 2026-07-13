from datetime import datetime, date
from pathlib import Path
import re
import unicodedata

from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.mitre.models import MitreAttack
from apps.sources.matching import sync_usecase_sources
from apps.usecases.models import UseCase


COLUMN_MAP = {
    "IDENTIFICADOR": "case_code",
    "CODIGO": "case_code",
    "GRUPO": "group_name",
    "DISPOSITIVO": "device",
    "FUENTES": "event_sources_raw",
    "TIPO": "case_type",
    "OBJETIVO2": "objective",
    "Tipo_bloqueo": "blocking_type",
    "NOMBRE NETWITNESS": "name",
    "RESPONSABLE": "owner_name",
    "Monitoreo": "monitoring",
    "status2": "status",
    "Fecha alta/ajuste": "created_or_adjusted_at",
    "Fecha puesta en producción": "production_date",
    "MITRE ATTACK": "mitre_attack_rel",
    "MITRE ATT&CK": "mitre_attack_rel",
    "Severidad": "severity",
    "Escalamiento": "escalation",
    "ENVIO.HO": "sent_to_ho",
    "HO": "ho_flag",
    "Última validación": "last_validation_date",
    "Fecha última validación": "last_validation_date",
    "last_validation_date": "last_validation_date",
}

DATE_FIELDS = {
    "created_or_adjusted_at",
    "production_date",
    "last_review_date",
    "next_review_date",
    "last_validation_date",
}

ATTACK_ID_RE = re.compile(r"\bT\d{4}(?:\.\d{3})?\b", re.IGNORECASE)
ATTACK_NUMERIC_ID_RE = re.compile(r"(?<![A-Z0-9])\d{4}(?:\.\d{3})?(?![A-Z0-9])", re.IGNORECASE)
D3FEND_HEADERS = {"D3FEND", "D3F3ND"}


def normalize_header(value):
    return "" if value is None else str(value).strip()


def normalize_text(value):
    return "" if value is None else str(value).strip()


def normalize_key(value):
    """Clave estable para comparar valores importados sin depender de acentos o formato."""
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


HEADER_ALIASES = {
    **{normalize_key(header): field for header, field in COLUMN_MAP.items()},
    "nombre": "name",
    "identificador": "case_code",
    "codigo": "case_code",
    "codigocaso": "case_code",
    "codigocasodeuso": "case_code",
    "nombrenetwitness": "name",
    "casodeuso": "name",
    "usecase": "name",
    "estado": "status",
    "status": "status",
    "objetivo": "objective",
    "tipobloqueo": "blocking_type",
    "fechaaltajuste": "created_or_adjusted_at",
    "fechapuestaenproduccion": "production_date",
    "fechaproduccion": "production_date",
    "mitre": "mitre_attack_rel",
    "attack": "mitre_attack_rel",
    "attck": "mitre_attack_rel",
    "mitreattack": "mitre_attack_rel",
    "mitreattck": "mitre_attack_rel",
    "mitreid": "mitre_attack_rel",
    "attackid": "mitre_attack_rel",
    "attackids": "mitre_attack_rel",
    "mitreattackid": "mitre_attack_rel",
    "mitreattackids": "mitre_attack_rel",
    "mitreattckid": "mitre_attack_rel",
    "mitreattckids": "mitre_attack_rel",
    "mitreattckrelacionado": "mitre_attack_rel",
    "mitreattackrelacionado": "mitre_attack_rel",
    "mitretecnicas": "mitre_attack_rel",
    "mitreattcktecnicas": "mitre_attack_rel",
    "mitreattacktecnicas": "mitre_attack_rel",
    "mitretechniques": "mitre_attack_rel",
    "tecnicasmitre": "mitre_attack_rel",
    "tecnicasattack": "mitre_attack_rel",
    "tecnicasattck": "mitre_attack_rel",
    "fuente": "event_sources_raw",
    "fuentes": "event_sources_raw",
    "fuenteseventos": "event_sources_raw",
    "fuentesdeeventos": "event_sources_raw",
    "eventsource": "event_sources_raw",
    "eventsources": "event_sources_raw",
    "ultimavalidacion": "last_validation_date",
    "fechaultimavalidacion": "last_validation_date",
}


def normalize_choice(value, choices, aliases=None):
    """
    Normaliza un valor de Excel contra los choices reales del modelo.

    Evita guardar variantes como "automático", "semiautomático" o "Sí"
    cuando el modelo espera exactamente "Automático", "Semiautomático" o "Sí".
    Si no reconoce el valor, lo devuelve sin modificar para no ocultar datos inesperados.
    """
    text = normalize_text(value)
    if not text:
        return ""

    aliases = aliases or {}
    key = normalize_key(text)

    if key in aliases:
        return aliases[key]

    valid_values = [choice_value for choice_value, _label in choices]
    valid_by_key = {normalize_key(choice_value): choice_value for choice_value in valid_values}

    return valid_by_key.get(key, text)


def normalize_status(value):
    aliases = {
        "testing": "Test",
        "test": "Test",
        "prueba": "Test",
        "qa": "Test",
        "prod": "Producción",
        "productivo": "Producción",
        "produccion": "Producción",
        "enproduccion": "Producción",
        "desarrollo": "Desarrollo",
        "dev": "Desarrollo",
        "baja": "Baja",
        "inactivo": "Baja",
        "disabled": "Baja",
        "propuesta": "Propuesta",
        "propuesto": "Propuesta",
        "rechazado": "Propuesta",
    }
    return normalize_choice(value, UseCase.STATUS_CHOICES, aliases)


def normalize_blocking_type(value):
    aliases = {
        "aut": "Automático",
        "auto": "Automático",
        "automatico": "Automático",
        "automatizado": "Automático",
        "manual": "Manual",
        "semi": "Semiautomático",
        "semiauto": "Semiautomático",
        "semiautomatico": "Semiautomático",
    }
    return normalize_choice(value, UseCase.BLOCKING_TYPE_CHOICES, aliases)


def normalize_severity(value):
    aliases = {
        "critical": "Critical",
        "critica": "Critical",
        "critico": "Critical",
        "alta": "High",
        "alto": "High",
        "high": "High",
        "media": "Medium",
        "medio": "Medium",
        "medium": "Medium",
        "baja": "Low",
        "bajo": "Low",
        "low": "Low",
    }
    return normalize_choice(value, UseCase.SEVERITY_CHOICES, aliases)


def normalize_escalation(value):
    aliases = {
        "irt": "IRT",
        "csirt": "IRT",
        "soc": "SOC",
        "otro": "Otro",
        "otros": "Otro",
        "n/a": "Otro",
        "na": "Otro",
    }
    return normalize_choice(value, UseCase.ESCALATION_CHOICES, aliases)


def normalize_yes_no(value):
    aliases = {
        "1": "Sí",
        "si": "Sí",
        "s": "Sí",
        "yes": "Sí",
        "y": "Sí",
        "true": "Sí",
        "x": "Sí",
        "0": "No",
        "no": "No",
        "n": "No",
        "false": "No",
    }
    return normalize_choice(value, UseCase.YES_NO_CHOICES, aliases)


def parse_date(value):
    if value in (None, ""):
        return None

    if isinstance(value, date):
        return value if not isinstance(value, datetime) else value.date()

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def extract_attack_ids(value):
    if not value:
        return []
    text = str(value)
    found = [item.upper() for item in ATTACK_ID_RE.findall(text)]
    found.extend(f"T{item}" for item in ATTACK_NUMERIC_ID_RE.findall(text))
    return sorted(set(found))


def extract_attack_ids_from_row(row):
    found = []
    for value in row:
        found.extend(extract_attack_ids(value))
    return sorted(set(found))


def resolve_attack_objects(attack_ids):
    if not attack_ids:
        return [], []

    attacks = list(MitreAttack.objects.filter(external_id__in=attack_ids).order_by("external_id"))
    found_ids = {attack.external_id.upper() for attack in attacks}
    missing_ids = [attack_id for attack_id in attack_ids if attack_id.upper() not in found_ids]
    return attacks, missing_ids


def find_existing_usecase(name, case_code=""):
    case_code = normalize_text(case_code)
    if case_code:
        instance = UseCase.objects.filter(case_code__iexact=case_code).first()
        if instance:
            return instance

    name = normalize_text(name)
    if not name:
        return None

    instance = UseCase.objects.filter(name__iexact=name).first()
    if instance:
        return instance

    name_key = normalize_key(name)
    for candidate in UseCase.objects.only("id", "name").iterator():
        if normalize_key(candidate.name) == name_key:
            return candidate
    return None


def sync_event_sources(usecase, raw_value):
    return sync_usecase_sources(
        usecase,
        raw_value,
        create_missing=True,
        defaults={"description": "Creada automaticamente desde import_usecases."},
    )


def validate_import_business_rules(payload, attack_ids):
    instance = UseCase(**payload)
    instance._clean_mitre_attack_ids = set(attack_ids)

    try:
        instance.clean()
        return []
    except ValidationError as exc:
        if hasattr(exc, "message_dict"):
            errors = []
            for field_errors in exc.message_dict.values():
                errors.extend(str(item) for item in field_errors)
            return errors
        return [str(item) for item in exc.messages]


class Command(BaseCommand):
    help = "Importa casos de uso desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Ruta al archivo Excel")
        parser.add_argument("--sheet", type=str, default=None, help="Nombre de la hoja")
        parser.add_argument(
            "--update",
            action="store_true",
            help="Actualiza registros existentes si encuentra el mismo nombre",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        sheet_name = options["sheet"]
        allow_update = options["update"]

        if not excel_path.exists():
            raise CommandError(f"No existe el archivo: {excel_path}")
        if excel_path.suffix.lower() != ".xlsx":
            raise CommandError("Solo se permiten archivos .xlsx. No se aceptan .xlsm ni otros formatos con macros.")

        wb = load_workbook(excel_path, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CommandError(
                    f"La hoja '{sheet_name}' no existe. Hojas disponibles: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("La hoja está vacía.")

        headers = [normalize_header(h) for h in rows[0]]
        d3fend_header_keys = {normalize_key(item) for item in D3FEND_HEADERS}
        ignored_d3fend_headers = [header for header in headers if normalize_key(header) in d3fend_header_keys]
        if ignored_d3fend_headers:
            self.stdout.write(self.style.WARNING(
                "La columna D3FEND del Excel será ignorada: D3FEND se infiere automáticamente desde MITRE ATT&CK."
            ))

        mapped_indexes = {}

        for idx, header in enumerate(headers):
            field_name = HEADER_ALIASES.get(normalize_key(header))
            if field_name:
                mapped_indexes[idx] = field_name

        if "name" not in mapped_indexes.values():
            raise CommandError("No se encontró la columna 'NOMBRE NETWITNESS'.")

        has_mitre_column = "mitre_attack_rel" in mapped_indexes.values()
        if not has_mitre_column:
            self.stdout.write(self.style.WARNING(
                "No se detectó una columna MITRE ATT&CK reconocida. "
                "Se buscarán IDs ATT&CK Txxxx en toda cada fila."
            ))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0
        warning_count = 0
        attack_assigned_count = 0

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                payload = {}
                attack_raw = ""
                sources_raw = None

                for idx, field_name in mapped_indexes.items():
                    raw_value = row[idx] if idx < len(row) else None

                    if field_name == "mitre_attack_rel":
                        attack_raw = normalize_text(raw_value)
                        continue
                    if field_name == "event_sources_raw":
                        sources_raw = normalize_text(raw_value)
                        continue

                    if field_name in DATE_FIELDS:
                        payload[field_name] = parse_date(raw_value)
                    else:
                        payload[field_name] = normalize_text(raw_value)

                payload["status"] = normalize_status(payload.get("status", ""))
                payload["blocking_type"] = normalize_blocking_type(payload.get("blocking_type", ""))
                payload["severity"] = normalize_severity(payload.get("severity", ""))
                payload["escalation"] = normalize_escalation(payload.get("escalation", ""))
                payload["sent_to_ho"] = normalize_yes_no(payload.get("sent_to_ho", ""))
                payload["ho_flag"] = normalize_yes_no(payload.get("ho_flag", ""))

                name = payload.get("name", "").strip()
                if not name:
                    skipped_count += 1
                    self.stdout.write(self.style.WARNING(f"Fila {row_num}: omitida por no tener nombre."))
                    continue
                if not payload.get("case_code"):
                    payload["case_code"] = name

                attack_ids = extract_attack_ids(attack_raw)
                if not attack_ids:
                    attack_ids = extract_attack_ids_from_row(row)
                    if attack_ids:
                        warning_count += 1
                        self.stdout.write(self.style.WARNING(
                            f"Fila {row_num}: MITRE detectado por busqueda en toda la fila -> {', '.join(attack_ids)}."
                        ))
                attack_objects, missing_attack_ids = resolve_attack_objects(attack_ids)
                resolved_attack_pk_ids = [attack.pk for attack in attack_objects]

                if attack_ids and missing_attack_ids:
                    warning_count += 1
                    self.stdout.write(self.style.WARNING(
                        f"Fila {row_num}: ATT&CK no encontrados en el catalogo -> {', '.join(missing_attack_ids)}."
                    ))

                business_warnings = validate_import_business_rules(payload, resolved_attack_pk_ids)
                if business_warnings:
                    warning_count += 1
                    self.stdout.write(self.style.WARNING(
                        f"Fila {row_num}: guardada con datos incompletos -> " + "; ".join(business_warnings)
                    ))

                instance = find_existing_usecase(name, payload.get("case_code", ""))

                if instance:
                    if not allow_update:
                        skipped_count += 1
                        self.stdout.write(self.style.WARNING(
                            f"Fila {row_num}: ya existe '{instance.name}', omitido. "
                            "Marca 'Actualizar existentes por nombre' para actualizar MITRE y datos."
                        ))
                        continue

                    for field, value in payload.items():
                        setattr(instance, field, value)
                    instance.save()
                    action = "actualizado"
                    updated_count += 1
                else:
                    instance = UseCase.objects.create(**payload)
                    action = "creado"
                    created_count += 1

                if hasattr(instance, "mitre_attacks"):
                    if allow_update:
                        instance.mitre_attacks.clear()
                    if attack_objects:
                        instance.mitre_attacks.add(*attack_objects)
                        attack_assigned_count += len(attack_objects)
                        self.stdout.write(
                            f"Fila {row_num}: MITRE asociado -> "
                            + ", ".join(attack.external_id for attack in attack_objects)
                        )

                instance.sync_d3fends_from_attacks()
                if attack_objects and not instance.d3fends.exists():
                    warning_count += 1
                    self.stdout.write(self.style.WARNING(
                        f"Fila {row_num}: ATT&CK cargado, pero no se infirio D3FEND. "
                        "Revisa que existan mappings D3FEND->ATT&CK cargados."
                    ))
                sync_event_sources(instance, sources_raw)

                self.stdout.write(self.style.SUCCESS(f"Fila {row_num}: {action} '{name}'"))

            except Exception as exc:
                error_count += 1
                self.stdout.write(self.style.ERROR(f"Fila {row_num}: error -> {exc}"))

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Importación finalizada"))
        self.stdout.write(f"Creados: {created_count}")
        self.stdout.write(f"Actualizados: {updated_count}")
        self.stdout.write(f"Omitidos: {skipped_count}")
        self.stdout.write(f"Errores: {error_count}")
        self.stdout.write(f"Advertencias: {warning_count}")
        self.stdout.write(f"MITRE asociados: {attack_assigned_count}")
