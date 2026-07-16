from datetime import date
from io import BytesIO, StringIO
import csv

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.lifecycle.models import LifecycleReview, LifecycleSettings
from apps.mitre.coverage_admin import build_coverage_admin_context
from apps.mitre.models import CoverageOverride, D3Fend, MitreAttack
from apps.sources.models import EventSource, UseCaseSource

from .forms import UseCaseForm
from .models import UseCase, UseCaseChangeLog, UseCaseEscalationOption, UseCaseRuleCondition


class UseCaseBusinessRuleTests(TestCase):
    def test_case_code_defaults_to_name_when_empty(self):
        usecase = UseCase.objects.create(name="Visible inventory name")

        self.assertEqual(usecase.case_code, "Visible inventory name")
        self.assertEqual(usecase.display_code, "Visible inventory name")

    def test_production_usecase_requires_mitre_mapping(self):
        usecase = UseCase(
            name="Suspicious process execution",
            status=UseCase.STATUS_PRODUCTION,
            production_date=date(2026, 1, 1),
        )
        usecase._clean_mitre_attack_ids = set()

        with self.assertRaises(ValidationError) as ctx:
            usecase.clean()

        self.assertIn("mitre_attacks", ctx.exception.message_dict)

    def test_sync_d3fends_from_attacks_sets_inferred_cache(self):
        attack = MitreAttack.objects.create(external_id="T1059", name="Command and Scripting Interpreter")
        d3fend = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        d3fend.related_attacks.add(attack)
        usecase = UseCase.objects.create(name="Command shell alert")
        usecase.mitre_attacks.add(attack)

        changed = usecase.sync_d3fends_from_attacks()

        self.assertTrue(changed)
        self.assertEqual(list(usecase.d3fends.order_by("id")), [d3fend])

    def test_d3fend_exclusion_removes_inferred_cache_for_usecase_only(self):
        attack = MitreAttack.objects.create(external_id="T1110", name="Brute Force")
        d3fend = D3Fend.objects.create(code="D3-ANAA", name="Account Authentication Analysis")
        d3fend.related_attacks.add(attack)
        usecase = UseCase.objects.create(name="VPN brute force alert")
        usecase.mitre_attacks.add(attack)
        usecase.sync_d3fends_from_attacks()
        self.assertEqual(list(usecase.d3fends.order_by("id")), [d3fend])
        usecase.d3fend_exclusions.add(d3fend)

        changed = usecase.sync_d3fends_from_attacks()

        self.assertTrue(changed)
        self.assertFalse(usecase.d3fends.exists())
        self.assertFalse(usecase.inferred_d3fends_queryset().exists())

    def test_parent_attack_does_not_infer_d3fend_from_subtechnique_mapping(self):
        parent = MitreAttack.objects.create(external_id="T1110", name="Brute Force")
        subtechnique = MitreAttack.objects.create(external_id="T1110.001", name="Password Guessing")
        d3fend = D3Fend.objects.create(code="D3-LAM", name="Local Account Monitoring")
        d3fend.related_attacks.add(subtechnique)
        usecase = UseCase.objects.create(name="Brute force alert")
        usecase.mitre_attacks.add(parent)

        changed = usecase.sync_d3fends_from_attacks()

        self.assertFalse(changed)
        self.assertFalse(usecase.d3fends.exists())

    def test_subtechnique_attack_does_not_infer_d3fend_from_parent_mapping(self):
        parent = MitreAttack.objects.create(external_id="T1110", name="Brute Force")
        subtechnique = MitreAttack.objects.create(external_id="T1110.003", name="Password Spraying")
        d3fend = D3Fend.objects.create(code="D3-ANAA", name="Account Authentication Analysis")
        d3fend.related_attacks.add(parent)
        usecase = UseCase.objects.create(name="Password spraying alert")
        usecase.mitre_attacks.add(subtechnique)

        changed = usecase.sync_d3fends_from_attacks()

        self.assertFalse(changed)
        self.assertFalse(usecase.d3fends.exists())

    def test_attack_family_expansion_does_not_cross_unrelated_prefixes(self):
        selected = MitreAttack.objects.create(external_id="T1110", name="Brute Force")
        unrelated = MitreAttack.objects.create(external_id="T1111", name="Unrelated")
        d3fend = D3Fend.objects.create(code="D3-OTHER", name="Other Defensive Technique")
        d3fend.related_attacks.add(unrelated)
        usecase = UseCase.objects.create(name="Brute force only")
        usecase.mitre_attacks.add(selected)

        changed = usecase.sync_d3fends_from_attacks()

        self.assertFalse(changed)
        self.assertFalse(usecase.d3fends.exists())

    def test_invalid_attack_ids_do_not_infer_every_d3fend(self):
        attack = MitreAttack.objects.create(external_id="T1059", name="Command and Scripting Interpreter")
        d3fend = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        d3fend.related_attacks.add(attack)

        inferred = UseCase.inferred_d3fends_for_attack_ids_queryset([999999])

        self.assertFalse(inferred.exists())

    def test_usecase_form_uses_escalation_catalog_and_hides_ho_flag(self):
        UseCaseEscalationOption.objects.create(name="NOC", position=40)

        form = UseCaseForm()
        escalation_values = [value for value, _ in form.fields["escalation"].choices]

        self.assertIn("NOC", escalation_values)
        self.assertIn('<option value="NOC">NOC</option>', str(form["escalation"]))
        self.assertNotIn("ho_flag", form.fields)


class SeedDemoDataCommandTests(TestCase):
    def test_seed_demo_data_creates_demo_catalog_users_and_cases(self):
        call_command("seed_demo_data", verbosity=0, stdout=StringIO())

        User = get_user_model()
        self.assertTrue(User.objects.filter(username="demo_admin", is_superuser=True).exists())
        self.assertTrue(User.objects.filter(username="demo_analyst").exists())
        self.assertGreaterEqual(MitreAttack.objects.filter(external_id__in=["T1059", "T1078", "T1110"]).count(), 3)
        self.assertGreaterEqual(D3Fend.objects.filter(code__startswith="D3-").count(), 5)
        self.assertGreaterEqual(UseCase.objects.filter(name__startswith="Demo - ").count(), 8)
        self.assertTrue(LifecycleReview.objects.filter(use_case__name__startswith="Demo - ").exists())
        self.assertTrue(CoverageOverride.objects.filter(reason__icontains="[demo]").exists())

    def test_seed_demo_data_is_idempotent(self):
        call_command("seed_demo_data", verbosity=0, stdout=StringIO())
        call_command("seed_demo_data", verbosity=0, stdout=StringIO())

        self.assertEqual(get_user_model().objects.filter(username="demo_admin").count(), 1)
        self.assertEqual(MitreAttack.objects.filter(external_id="T1059").count(), 1)
        self.assertEqual(UseCase.objects.filter(name="Demo - PowerShell suspicious execution").count(), 1)
        self.assertTrue(LifecycleSettings.objects.get(name="Demo lifecycle").is_active)


class UseCasePermissionTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin_group = Group.objects.create(name="Admin")
        self.analyst_group = Group.objects.create(name="Analyst")
        self.readonly_group = Group.objects.create(name="ReadOnly")

        self.admin = User.objects.create_user("admin", password="pass")
        self.admin.groups.add(self.admin_group)
        self.analyst = User.objects.create_user("analyst", password="pass")
        self.analyst.groups.add(self.analyst_group)
        self.other_analyst = User.objects.create_user("other", password="pass")
        self.other_analyst.groups.add(self.analyst_group)
        self.readonly = User.objects.create_user("readonly", password="pass")
        self.readonly.groups.add(self.readonly_group)

        self.attack = MitreAttack.objects.create(external_id="T1059", name="Command and Scripting Interpreter")
        self.owned_usecase = self._create_production_usecase("Owned use case", self.analyst)
        self.other_usecase = self._create_production_usecase("Other use case", self.other_analyst)

    def _create_production_usecase(self, name, creator):
        usecase = UseCase.objects.create(
            name=name,
            owner_name=creator.username,
            created_by=creator,
            status=UseCase.STATUS_PRODUCTION,
            production_date=date(2026, 1, 1),
            severity="Low",
            is_enabled=True,
        )
        usecase.mitre_attacks.add(self.attack)
        return usecase

    def _bulk_payload(self, *usecases):
        data = {"changed_ids": ",".join(str(usecase.pk) for usecase in usecases)}
        for usecase in usecases:
            data.update({
                f"owner_name_{usecase.pk}": usecase.owner_name,
                f"severity_{usecase.pk}": "High",
                f"last_validation_date_{usecase.pk}": "",
                f"is_enabled_{usecase.pk}": "on",
                f"status_{usecase.pk}": usecase.status,
                f"validation_status_{usecase.pk}": usecase.validation_status,
                f"validation_result_{usecase.pk}": usecase.validation_result,
                f"mitre_attack_ids_{usecase.pk}": str(self.attack.pk),
            })
        return data

    def test_readonly_cannot_access_usecase_inventory(self):
        self.client.login(username="readonly", password="pass")

        response = self.client.get(reverse("usecase_list"))

        self.assertEqual(response.status_code, 403)

    def test_legacy_usecase_routes_redirect_by_default(self):
        response = self.client.get("/usecases/attack-matrix/")

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("attack_matrix"), response["Location"])

    @override_settings(ENABLE_LEGACY_USECASE_REDIRECTS=False)
    def test_legacy_usecase_routes_can_be_disabled(self):
        response = self.client.get("/usecases/attack-matrix/")

        self.assertEqual(response.status_code, 404)

    def test_readonly_cannot_export_csv_or_pdf(self):
        self.client.login(username="readonly", password="pass")

        csv_response = self.client.get(reverse("export_usecases_csv"))
        full_csv_response = self.client.get(reverse("export_usecases_full_csv"))
        pdf_response = self.client.get(reverse("dashboard_pdf_export"))

        self.assertEqual(csv_response.status_code, 403)
        self.assertEqual(full_csv_response.status_code, 403)
        self.assertEqual(pdf_response.status_code, 403)

    def test_analyst_can_export_dashboard_pdf(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("dashboard_pdf_export"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("dashboard-soc-", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_csv_export_respects_device_filter(self):
        self.owned_usecase.device = "EDR"
        self.owned_usecase.save()
        self.other_usecase.device = "SIEM"
        self.other_usecase.save()
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("export_usecases_csv"), {"device": "EDR"})
        content = response.content.decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Owned use case", content)
        self.assertNotIn("Other use case", content)

    def test_csv_export_only_includes_production_usecases(self):
        draft = UseCase.objects.create(
            name="Draft only use case",
            owner_name=self.analyst.username,
            created_by=self.analyst,
            status=UseCase.STATUS_DEVELOPMENT,
            severity="Low",
            is_enabled=True,
        )
        draft.mitre_attacks.add(self.attack)
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("export_usecases_csv"))
        content = response.content.decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Owned use case", content)
        self.assertIn("Other use case", content)
        self.assertNotIn("Draft only use case", content)

    def test_inventory_list_uses_exact_attack_d3fend_inference(self):
        parent = MitreAttack.objects.create(external_id="T1110", name="Brute Force")
        subtechnique = MitreAttack.objects.create(external_id="T1110.001", name="Password Guessing")
        d3fend = D3Fend.objects.create(code="D3-LAM", name="Local Account Monitoring")
        d3fend.related_attacks.add(subtechnique)
        usecase = self._create_production_usecase("Brute family inventory case", self.analyst)
        usecase.mitre_attacks.set([parent])
        usecase.sync_d3fends_from_attacks()
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("usecase_list"), {"q": "Brute family inventory case"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "D3FEND 0")

    def test_analyst_can_export_usecases_xlsx(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("export_usecases_xlsx"))
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(sheet["A1"].value, "IDENTIFICADOR")
        self.assertEqual(sheet["G1"].value, "NOMBRE NETWITNESS")
        self.assertIn("Owned use case", [cell.value for cell in sheet["G"]])

    def test_analyst_can_export_full_inventory_csv(self):
        d3fend = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        d3fend.related_attacks.add(self.attack)
        self.owned_usecase.full_rule_text = "SELECT * FROM Event;"
        self.owned_usecase.save()
        self.owned_usecase.sync_d3fends_from_attacks()
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("export_usecases_full_csv"))
        content = response.content.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(content)))

        self.assertEqual(response.status_code, 200)
        self.assertIn("D3FEND_EXCLUIDO", rows[0])
        self.assertIn("D3FEND_INFERIDO", rows[0])
        self.assertIn("Regla completa", rows[0])
        owned_row = next(row for row in rows if row["NOMBRE NETWITNESS"] == "Owned use case")
        self.assertIn("T1059", owned_row["MITRE ATT&CK"])
        self.assertIn("D3-PSEP", owned_row["D3FEND_INFERIDO"])
        self.assertEqual(owned_row["Regla completa"], "SELECT * FROM Event;")

    def test_analyst_can_download_import_template(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.get(reverse("download_usecase_import_template"))
        workbook = load_workbook(BytesIO(response.content))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(workbook.active["A1"].value, "IDENTIFICADOR")
        self.assertEqual(workbook.active["G1"].value, "NOMBRE NETWITNESS")

    def test_analyst_can_import_full_inventory_csv(self):
        d3fend = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        d3fend.related_attacks.add(self.attack)
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "IDENTIFICADOR",
            "NOMBRE NETWITNESS",
            "status2",
            "Fecha puesta en produccion",
            "MITRE ATT&CK",
            "D3FEND_EXCLUIDO",
            "FUENTES",
            "Habilitado",
            "Regla completa",
            "Descripcion funcional",
        ])
        writer.writerow([
            "CSV-001",
            "Imported CSV use case",
            "Produccion",
            "2026-01-02",
            "T1059 - Command and Scripting Interpreter",
            "D3-PSEP - Process Spawn Analysis",
            "SRC-CSV - CSV Source",
            "Si",
            "SELECT * FROM Event;",
            "Detecta comportamiento de prueba.",
        ])
        upload = SimpleUploadedFile(
            "inventario.csv",
            buffer.getvalue().encode("utf-8-sig"),
            content_type="text/csv",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_csv"), {"csv_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(case_code="CSV-001")
        self.assertEqual(imported.name, "Imported CSV use case")
        self.assertEqual(imported.status, UseCase.STATUS_PRODUCTION)
        self.assertEqual(imported.full_rule_text, "SELECT * FROM Event;")
        self.assertEqual(imported.functional_description, "Detecta comportamiento de prueba.")
        self.assertEqual(list(imported.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(imported.d3fend_exclusions.values_list("code", flat=True)), ["D3-PSEP"])
        self.assertFalse(imported.d3fends.exists())
        self.assertTrue(EventSource.objects.filter(code="SRC-CSV", name="CSV Source").exists())
        self.assertEqual(list(imported.source_links.values_list("source__code", flat=True)), ["SRC-CSV"])

    def test_full_inventory_csv_ignores_d3fend_inferred_column(self):
        curated = D3Fend.objects.create(code="D3-NTSA", name="Network Traffic Signature Analysis")
        official = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        official.related_attacks.add(self.attack)
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "IDENTIFICADOR",
            "NOMBRE NETWITNESS",
            "status2",
            "MITRE ATT&CK",
            "D3FEND_INFERIDO",
        ])
        writer.writerow([
            "CSV-002",
            "Imported curated D3FEND case",
            "Produccion",
            "T1059 - Command and Scripting Interpreter",
            "D3-NTSA - Network Traffic Signature Analysis",
        ])
        upload = SimpleUploadedFile(
            "inventario.csv",
            buffer.getvalue().encode("utf-8-sig"),
            content_type="text/csv",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_csv"), {"csv_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(case_code="CSV-002")
        self.assertEqual(list(imported.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(imported.d3fends.values_list("code", flat=True)), [official.code])
        self.assertNotIn(curated, imported.d3fends.all())
        output = self.client.session.get("last_usecase_import_output", "")
        self.assertNotIn("D3FEND curados desde CSV", output)
        self.assertNotIn("no se infirio D3FEND", output)

    def test_analyst_can_import_usecases_excel(self):
        d3fend = D3Fend.objects.create(code="D3-PSEP", name="Process Spawn Analysis")
        d3fend.related_attacks.add(self.attack)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "GRUPO",
            "DISPOSITIVO",
            "TIPO",
            "OBJETIVO2",
            "Tipo_bloqueo",
            "NOMBRE NETWITNESS",
            "RESPONSABLE",
            "Monitoreo",
            "status2",
            "Fecha alta/ajuste",
            "Fecha puesta en producción",
            "MITRE ATT&CK",
            "Severidad",
            "Escalamiento",
            "ENVIO.HO",
            "HO",
            "FUENTES",
            "Fecha última validación",
        ])
        sheet.append([
            "SOC",
            "SIEM",
            "Correlation",
            "Detect test import",
            "Manual",
            "Imported Excel use case",
            "analyst",
            "24x7",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 1),
            date(2026, 1, 2),
            "T1059 - Command and Scripting Interpreter",
            "High",
            "SOC",
            "No",
            "No",
            "SRC-EDR - Endpoint EDR; CloudTrail",
            date(2026, 1, 3),
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported Excel use case")
        self.assertEqual(imported.case_code, "Imported Excel use case")
        self.assertEqual(imported.status, UseCase.STATUS_PRODUCTION)
        self.assertEqual(imported.device, "SIEM")
        self.assertEqual(list(imported.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(imported.d3fends.values_list("code", flat=True)), ["D3-PSEP"])
        self.assertTrue(EventSource.objects.filter(code="SRC-EDR", name="Endpoint EDR").exists())
        self.assertTrue(EventSource.objects.filter(name="CloudTrail").exists())
        self.assertEqual(
            set(imported.source_links.values_list("source__name", flat=True)),
            {"Endpoint EDR", "CloudTrail"},
        )

    def test_import_accepts_mitre_header_variants_and_infers_d3fend(self):
        d3fend = D3Fend.objects.create(code="D3-TEST", name="Test Defensive Technique")
        d3fend.related_attacks.add(self.attack)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Nombre NetWitness",
            "Estado",
            "Fecha producción",
            "MITRE ATT&CK relacionado",
        ])
        sheet.append([
            "Imported variant attack header",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 2),
            "T1059",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported variant attack header")
        self.assertEqual(list(imported.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(imported.d3fends.values_list("code", flat=True)), ["D3-TEST"])

    def test_import_accepts_real_mitre_tecnicas_header(self):
        second_attack = MitreAttack.objects.create(external_id="T1595", name="Active Scanning")
        d3fend = D3Fend.objects.create(code="D3-CAA", name="Connection Attempt Analysis")
        d3fend.related_attacks.add(self.attack, second_attack)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "NOMBRE NETWITNESS",
            "status2",
            "Fecha puesta en producción",
            "MITRE Tecnicas",
        ])
        sheet.append([
            "Imported real mitre header",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 2),
            "T1595, T1059",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported real mitre header")
        self.assertEqual(
            list(imported.mitre_attacks.order_by("external_id").values_list("external_id", flat=True)),
            ["T1059", "T1595"],
        )
        self.assertEqual(list(imported.d3fends.values_list("code", flat=True)), ["D3-CAA"])
        output = self.client.session.get("last_usecase_import_output", "")
        self.assertNotIn("No se detectó una columna MITRE", output)
        self.assertIn("MITRE asociado -> T1059, T1595", output)

    def test_import_saves_production_row_with_warning_when_attack_id_is_not_in_catalog(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "NOMBRE NETWITNESS",
            "status2",
            "Fecha puesta en producción",
            "MITRE ATT&CK",
        ])
        sheet.append([
            "Imported missing attack",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 2),
            "T9999",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported missing attack")
        self.assertFalse(imported.mitre_attacks.exists())
        output = self.client.session.get("last_usecase_import_output", "")
        self.assertIn("ATT&CK no encontrados en el catalogo -> T9999", output)
        self.assertIn("guardada con datos incompletos", output)

    def test_import_update_matches_existing_name_flexibly_and_updates_mitre(self):
        d3fend = D3Fend.objects.create(code="D3-UPDATE", name="Update Defensive Technique")
        d3fend.related_attacks.add(self.attack)
        existing = UseCase.objects.create(
            name="Imported Existing Case",
            status=UseCase.STATUS_PRODUCTION,
            production_date=date(2026, 1, 1),
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "NOMBRE NETWITNESS",
            "status2",
            "Fecha puesta en producción",
            "MITRE ATT&CK",
        ])
        sheet.append([
            " imported   existing case ",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 2),
            "T1059",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(
            reverse("import_usecases_excel"),
            {"excel_file": upload, "update_existing": "on"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(UseCase.objects.filter(name__icontains="existing case").count(), 1)
        existing.refresh_from_db()
        self.assertEqual(existing.name, "imported   existing case")
        self.assertEqual(list(existing.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(existing.d3fends.values_list("code", flat=True)), ["D3-UPDATE"])
        output = self.client.session.get("last_usecase_import_output", "")
        self.assertIn("MITRE asociado -> T1059", output)

    def test_import_falls_back_to_scanning_row_for_attack_ids(self):
        d3fend = D3Fend.objects.create(code="D3-FALLBACK", name="Fallback Defensive Technique")
        d3fend.related_attacks.add(self.attack)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "NOMBRE NETWITNESS",
            "status2",
            "Fecha puesta en producción",
            "Referencia de framework",
        ])
        sheet.append([
            "Imported attack fallback",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 2),
            "MITRE technique T1059",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported attack fallback")
        self.assertEqual(list(imported.mitre_attacks.values_list("external_id", flat=True)), ["T1059"])
        self.assertEqual(list(imported.d3fends.values_list("code", flat=True)), ["D3-FALLBACK"])
        output = self.client.session.get("last_usecase_import_output", "")
        self.assertIn("MITRE detectado por busqueda en toda la fila -> T1059", output)
        self.assertIn("MITRE asociado -> T1059", output)

    def test_import_without_fuentes_does_not_treat_device_as_event_source(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "GRUPO",
            "DISPOSITIVO",
            "TIPO",
            "OBJETIVO2",
            "Tipo_bloqueo",
            "NOMBRE NETWITNESS",
            "RESPONSABLE",
            "Monitoreo",
            "status2",
            "Fecha alta/ajuste",
            "Fecha puesta en producción",
            "MITRE ATT&CK",
            "Severidad",
            "Escalamiento",
            "ENVIO.HO",
            "HO",
            "Fecha última validación",
        ])
        sheet.append([
            "SOC",
            "Legacy FW",
            "Correlation",
            "Detect test import without source",
            "Manual",
            "Imported without source",
            "analyst",
            "24x7",
            UseCase.STATUS_PRODUCTION,
            date(2026, 1, 1),
            date(2026, 1, 2),
            "T1059 - Command and Scripting Interpreter",
            "High",
            "SOC",
            "No",
            "No",
            date(2026, 1, 3),
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "usecases.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        imported = UseCase.objects.get(name="Imported without source")
        self.assertEqual(imported.device, "Legacy FW")
        self.assertFalse(imported.source_links.exists())
        self.assertFalse(EventSource.objects.filter(name="Legacy FW").exists())

    def test_import_rejects_macro_enabled_workbook_extension(self):
        upload = SimpleUploadedFile(
            "usecases.xlsm",
            b"not-a-real-workbook",
            content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        )
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("import_usecases_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(UseCase.objects.filter(name="usecases.xlsm").exists())

    def test_readonly_cannot_import_or_export_xlsx(self):
        self.client.login(username="readonly", password="pass")

        xlsx_response = self.client.get(reverse("export_usecases_xlsx"))
        template_response = self.client.get(reverse("download_usecase_import_template"))
        import_response = self.client.get(reverse("import_usecases_excel"))
        csv_import_response = self.client.post(reverse("import_usecases_csv"))

        self.assertEqual(xlsx_response.status_code, 403)
        self.assertEqual(template_response.status_code, 403)
        self.assertEqual(import_response.status_code, 403)
        self.assertEqual(csv_import_response.status_code, 403)

    def test_analyst_bulk_update_only_changes_owned_usecases(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.post(
            reverse("usecase_bulk_update"),
            self._bulk_payload(self.owned_usecase, self.other_usecase),
        )

        self.assertEqual(response.status_code, 302)
        self.owned_usecase.refresh_from_db()
        self.other_usecase.refresh_from_db()
        self.assertEqual(self.owned_usecase.severity, "High")
        self.assertEqual(self.other_usecase.severity, "Low")

    def test_admin_bulk_update_can_change_any_usecase(self):
        self.client.login(username="admin", password="pass")

        response = self.client.post(
            reverse("usecase_bulk_update"),
            self._bulk_payload(self.other_usecase),
        )

        self.assertEqual(response.status_code, 302)
        self.other_usecase.refresh_from_db()
        self.assertEqual(self.other_usecase.severity, "High")

    def test_analyst_can_save_rule_conditions_and_full_rule(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("usecase_edit", args=[self.owned_usecase.pk]), {
            "group_name": self.owned_usecase.group_name,
            "device": self.owned_usecase.device,
            "case_type": self.owned_usecase.case_type,
            "objective": "Detect privileged group changes.",
            "blocking_type": "",
            "name": self.owned_usecase.name,
            "owner_name": self.owned_usecase.owner_name,
            "monitoring": self.owned_usecase.monitoring,
            "status": self.owned_usecase.status,
            "created_or_adjusted_at": "",
            "production_date": "2026-01-01",
            "mitre_attacks": [str(self.attack.pk)],
            "severity": self.owned_usecase.severity,
            "escalation": "",
            "sent_to_ho": "",
            "ho_flag": "",
            "last_validation_date": "",
            "validation_status": self.owned_usecase.validation_status,
            "validation_result": self.owned_usecase.validation_result,
            "is_enabled": "on",
            "disabled_reason": "",
            "comments": "Reviewable logic.",
            "full_rule_text": "SELECT * FROM Event WHERE source = 'Active Directory';",
            "functional_description": "Controla altas de usuarios privilegiados.",
            "event_sources": [],
            "conditions-TOTAL_FORMS": "2",
            "conditions-INITIAL_FORMS": "0",
            "conditions-MIN_NUM_FORMS": "0",
            "conditions-MAX_NUM_FORMS": "1000",
            "conditions-0-position": "1",
            "conditions-0-condition_type": UseCaseRuleCondition.TYPE_INCLUDE,
            "conditions-0-field_name": "source",
            "conditions-0-operator": UseCaseRuleCondition.OP_EQUALS,
            "conditions-0-value": "Active Directory",
            "conditions-0-use_case": str(self.owned_usecase.pk),
            "conditions-1-position": "2",
            "conditions-1-condition_type": UseCaseRuleCondition.TYPE_EXCLUDE,
            "conditions-1-field_name": "environment",
            "conditions-1-operator": UseCaseRuleCondition.OP_EQUALS,
            "conditions-1-value": "laboratorio",
            "conditions-1-use_case": str(self.owned_usecase.pk),
        })

        self.assertEqual(response.status_code, 302)
        self.owned_usecase.refresh_from_db()
        self.assertEqual(self.owned_usecase.rule_conditions.count(), 2)
        self.assertEqual(self.owned_usecase.full_rule_text, "SELECT * FROM Event WHERE source = 'Active Directory';")
        self.assertTrue(
            UseCaseChangeLog.objects.filter(
                use_case=self.owned_usecase,
                field_name="rule_conditions",
                new_value__icontains="Active Directory",
            ).exists()
        )

        detail_response = self.client.get(reverse("usecase_detail", args=[self.owned_usecase.pk]))
        self.assertContains(detail_response, "Regla y condiciones")
        self.assertContains(detail_response, "Active Directory")
        self.assertContains(detail_response, "laboratorio")
        self.assertContains(detail_response, "Controla altas de usuarios privilegiados.")

    def test_usecase_form_accepts_multiple_groups_and_devices(self):
        event_source = EventSource.objects.create(code="SRC-FW", name="Firewall Events")
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("usecase_edit", args=[self.owned_usecase.pk]), {
            "group_name": "Endpoints; Perimetral\nCloud",
            "device": "EDR, Firewall; CASB",
            "case_type": self.owned_usecase.case_type,
            "objective": self.owned_usecase.objective,
            "blocking_type": "",
            "name": self.owned_usecase.name,
            "owner_name": self.owned_usecase.owner_name,
            "monitoring": self.owned_usecase.monitoring,
            "status": self.owned_usecase.status,
            "created_or_adjusted_at": "",
            "production_date": "2026-01-01",
            "mitre_attacks": [str(self.attack.pk)],
            "severity": self.owned_usecase.severity,
            "escalation": "",
            "sent_to_ho": "",
            "ho_flag": "",
            "last_validation_date": "",
            "validation_status": self.owned_usecase.validation_status,
            "validation_result": self.owned_usecase.validation_result,
            "is_enabled": "on",
            "disabled_reason": "",
            "comments": "",
            "full_rule_text": "",
            "functional_description": "",
            "event_sources": [str(event_source.pk)],
            "conditions-TOTAL_FORMS": "0",
            "conditions-INITIAL_FORMS": "0",
            "conditions-MIN_NUM_FORMS": "0",
            "conditions-MAX_NUM_FORMS": "1000",
        })

        self.assertEqual(response.status_code, 302)
        self.owned_usecase.refresh_from_db()
        self.assertEqual(self.owned_usecase.group_name, "Endpoints, Perimetral, Cloud")
        self.assertEqual(self.owned_usecase.device, "EDR, Firewall, CASB")
        self.assertEqual(list(self.owned_usecase.source_links.values_list("source_id", flat=True)), [event_source.pk])

        list_response = self.client.get(reverse("usecase_list"), {"device": "Firewall"})
        self.assertContains(list_response, self.owned_usecase.name)

        edit_response = self.client.get(reverse("usecase_edit", args=[self.owned_usecase.pk]))
        self.assertContains(edit_response, "data-options")
        self.assertContains(edit_response, "data-multi-select")
        self.assertContains(edit_response, "Perimetral")
        self.assertContains(edit_response, "Firewall")
        self.assertContains(edit_response, "SRC-FW - Firewall Events")

    def test_coverage_override_requires_reason_for_fulfilled_status(self):
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("coverage_override_update"), {
            "framework": CoverageOverride.FRAMEWORK_ATTACK,
            "object_type": CoverageOverride.OBJECT_TECHNIQUE,
            "object_key": self.attack.external_id,
            "object_name": self.attack.name,
            "status": CoverageOverride.STATUS_FULFILLED,
            "reason": "",
            "default_enabled": "1",
        })

        self.assertEqual(response.status_code, 302)
        self.assertFalse(CoverageOverride.objects.exists())

    def test_coverage_admin_context_searches_tactics_by_child_technique(self):
        context = build_coverage_admin_context({
            "tab": CoverageOverride.FRAMEWORK_ATTACK,
            "scope": CoverageOverride.OBJECT_TACTIC,
            "q": "T1059",
        })

        self.assertEqual(context["tab"], CoverageOverride.FRAMEWORK_ATTACK)
        self.assertEqual(context["scope"], CoverageOverride.OBJECT_TACTIC)
        self.assertEqual(len(context["rows"]), 1)

    def test_lifecycle_owner_can_mark_review_done(self):
        self.owned_usecase.lifecycle_control_owner = self.analyst
        self.owned_usecase.save()
        source = EventSource.objects.create(name="EDR", source_type=EventSource.TYPE_EDR)
        UseCaseSource.objects.create(use_case=self.owned_usecase, source=source)
        self.client.login(username="analyst", password="pass")

        response = self.client.post(reverse("lifecycle_mark_done", args=[self.owned_usecase.pk]), {
            "validation_result": LifecycleReview.RESULT_CURRENT,
            "logic_valid": "on",
            "sources_active": "on",
            "event_ids_valid": "on",
            "fields_exist": "on",
            "trigger_count": "0",
            "notes": "Control funcional verificado.",
        })

        self.assertEqual(response.status_code, 302)
        self.owned_usecase.refresh_from_db()
        self.assertEqual(self.owned_usecase.validation_status, UseCase.VALIDATION_STATUS_FINISHED)
        self.assertEqual(self.owned_usecase.validation_result, UseCase.VALIDATION_RESULT_OK)
        self.assertEqual(LifecycleReview.objects.filter(use_case=self.owned_usecase).count(), 1)

    def test_non_owner_cannot_mark_lifecycle_review_done(self):
        self.owned_usecase.lifecycle_control_owner = self.analyst
        self.owned_usecase.save()
        self.client.login(username="other", password="pass")

        response = self.client.post(reverse("lifecycle_mark_done", args=[self.owned_usecase.pk]))

        self.assertEqual(response.status_code, 403)
        self.assertFalse(LifecycleReview.objects.filter(use_case=self.owned_usecase).exists())

    def test_admin_can_delete_usecase(self):
        self.client.login(username="admin", password="pass")

        response = self.client.post(reverse("usecase_delete", args=[self.other_usecase.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(UseCase.objects.filter(pk=self.other_usecase.pk).exists())



